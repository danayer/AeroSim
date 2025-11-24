"""
Менеджер для экспорта данных симуляции в различные форматы
"""

import json
import csv
from datetime import datetime
from typing import Dict, Any, List
from pathlib import Path

# Импорты для Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Импорты для PDF
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfgen import canvas
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ExportManager:
    """Управление экспортом данных в различные форматы"""
    
    _fonts_registered = False
    
    def __init__(self):
        self.export_formats = {
            'csv': self.export_csv,
            'json': self.export_json,
            'xlsx': self.export_xlsx,
        }
        self._register_fonts()
    
    @staticmethod
    def _register_fonts():
        """Зарегистрировать шрифты для поддержки кириллицы"""
        if not REPORTLAB_AVAILABLE or ExportManager._fonts_registered:
            return
        
        try:
            # Попробуем найти системный шрифт DejaVuSans, поддерживающий кириллицу
            import os
            
            # Возможные пути к шрифтам на различных Linux системах
            font_paths = [
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
            ]
            
            font_found = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('CustomFont', font_path))
                        pdfmetrics.registerFont(TTFont('CustomFontBold', font_path))
                        ExportManager._fonts_registered = True
                        font_found = True
                        break
                    except Exception:
                        continue
            
            if not font_found:
                # Используем встроенный шрифт если системный не найден
                ExportManager._fonts_registered = True
        except Exception:
            pass
    
    @staticmethod
    def _prepare_data(stats: Dict) -> Dict:
        """Подготовить данные для экспорта (рекурсивно преобразовать объекты)"""
        prepared = {}
        for key, value in stats.items():
            if isinstance(value, dict):
                prepared[key] = ExportManager._prepare_data(value)
            elif isinstance(value, (list, tuple)):
                prepared[key] = [
                    ExportManager._prepare_data(item) if isinstance(item, dict) else item
                    for item in value
                ]
            elif hasattr(value, '__dict__'):
                # Преобразовать объект в словарь
                prepared[key] = ExportManager._prepare_data(value.__dict__)
            else:
                prepared[key] = value
        return prepared
    
    def export_json(self, stats: Dict, file_path: str) -> bool:
        """
        Экспортировать данные в JSON формате
        
        Args:
            stats: Словарь статистики
            file_path: Путь для сохранения
            
        Returns:
            bool: Успешность операции
        """
        try:
            prepared_data = self._prepare_data(stats)
            
            # Добавить мета-информацию
            export_data = {
                'export_time': datetime.now().isoformat(),
                'export_format': 'JSON',
                'simulation_data': prepared_data,
                'metadata': {
                    'total_events': prepared_data.get('total_events_processed', 0),
                    'total_aircraft': prepared_data.get('total_aircraft', 0),
                    'total_passengers': prepared_data.get('total_passengers', 0),
                    'simulation_time': prepared_data.get('simulation_time', 0),
                }
            }
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False)
            
            return True
        except Exception as e:
            print(f"Ошибка при экспорте JSON: {e}")
            return False
    
    def export_csv(self, stats: Dict, file_path: str) -> bool:
        """
        Экспортировать данные в CSV формате
        
        Args:
            stats: Словарь статистики
            file_path: Путь для сохранения
            
        Returns:
            bool: Успешность операции
        """
        try:
            prepared_data = self._prepare_data(stats)
            
            # Создать плоскую структуру данных для CSV
            flat_data = self._flatten_dict(prepared_data)
            
            if not flat_data:
                return False
            
            # Определить заголовки
            headers = list(flat_data.keys())
            
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerow(flat_data)
            
            return True
        except Exception as e:
            print(f"Ошибка при экспорте CSV: {e}")
            return False
    
    def export_xlsx(self, stats: Dict, file_path: str) -> bool:
        """
        Экспортировать данные в XLSX формате
        
        Args:
            stats: Словарь статистики
            file_path: Путь для сохранения
            
        Returns:
            bool: Успешность операции
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            prepared_data = self._prepare_data(stats)
            
            # Создать новую рабочую книгу
            wb = openpyxl.Workbook()
            
            # === ЛИСТ 1: Основная статистика ===
            ws_main = wb.active
            ws_main.title = "Основная статистика"
            self._create_main_sheet(ws_main, prepared_data)
            
            # === ЛИСТ 2: Экономика ===
            if 'airport_economics' in prepared_data:
                ws_econ = wb.create_sheet("Экономика")
                self._create_economics_sheet(ws_econ, prepared_data['airport_economics'])
            
            # === ЛИСТ 3: Мониторинг ===
            ws_monitor = wb.create_sheet("Мониторинг")
            self._create_monitoring_sheet(ws_monitor, prepared_data)
            
            # === ЛИСТ 4: Детали ===
            ws_details = wb.create_sheet("Детали")
            self._create_details_sheet(ws_details, prepared_data)
            
            wb.save(file_path)
            return True
        except ImportError:
            print("Ошибка: Требуется установка openpyxl (pip install openpyxl)")
            return False
        except Exception as e:
            print(f"Ошибка при экспорте XLSX: {e}")
            return False
    
    @staticmethod
    def _flatten_dict(data: Dict, parent_key: str = '', sep: str = '_') -> List[Dict]:
        """Преобразовать вложенный словарь в плоскую структуру"""
        flat_data = {}
        
        for key, value in data.items():
            new_key = f"{parent_key}{sep}{key}" if parent_key else key
            
            if isinstance(value, dict):
                # Рекурсивно разворачиваем вложенные словари
                nested = ExportManager._flatten_dict(value, new_key, sep=sep)
                flat_data.update(nested)
            elif isinstance(value, (list, tuple)):
                flat_data[new_key] = str(value)
            else:
                flat_data[new_key] = value
        
        return flat_data
    
    @staticmethod
    def _create_main_sheet(ws, data: Dict):
        """Создать лист основной статистики"""
        # Заголовок
        ws['A1'] = "ОСНОВНАЯ СТАТИСТИКА АЭРОПОРТА"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Подзаголовок
        ws['A2'] = f"Дата экспорта: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10, italic=True)
        
        row = 4
        
        # Ключевые метрики
        metrics = [
            ('Всего обработано событий', data.get('total_events_processed', 0)),
            ('Всего самолётов', data.get('total_aircraft', 0)),
            ('Всего пассажиров', data.get('total_passengers', 0)),
            ('Время симуляции (сек)', round(data.get('simulation_time', 0), 2)),
            ('Пропускная способность (самолётов/ч)', round(data.get('throughput', 0), 2)),
            ('Среднее время обслуживания (сек)', round(data.get('avg_service_time', 0), 2)),
            ('Использование ВПП (%)', round(data.get('runway_utilization', 0), 2)),
            ('Использование гейтов (%)', round(data.get('gate_utilization', 0), 2)),
            ('Среднее использование (%)', round(data.get('average_utilization', 0), 2)),
            ('Режим операции', data.get('mode', 'Неизвестно')),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
    
    @staticmethod
    def _create_economics_sheet(ws, economics_data: Dict):
        """Создать лист экономики"""
        # Заголовок
        ws['A1'] = "ЭКОНОМИЧЕСКАЯ СТАТИСТИКА"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Финансовые показатели
        financial_metrics = [
            ('Всего рейсов', economics_data.get('total_flights', 0)),
            ('Местные рейсы', economics_data.get('commuter_flights', 0)),
            ('Международные рейсы', economics_data.get('international_flights', 0)),
            ('', None),  # Пустая строка
            ('ДОХОД', None),
            ('Общий доход', f"${economics_data.get('total_revenue', 0):,.2f}"),
            ('Доход First Class', f"${economics_data.get('first_class_revenue', 0):,.2f}"),
            ('Доход Coach', f"${economics_data.get('coach_revenue', 0):,.2f}"),
            ('', None),
            ('РАСХОДЫ И ПРИБЫЛЬ', None),
            ('Общие расходы', f"${economics_data.get('total_costs', 0):,.2f}"),
            ('Общая прибыль', f"${economics_data.get('total_profit', 0):,.2f}"),
            ('ROI (%)', f"{economics_data.get('roi_percentage', 0):.2f}%"),
            ('', None),
            ('ПАССАЖИРЫ', None),
            ('Всего обслужено пассажиров', economics_data.get('total_passengers_served', 0)),
            ('First Class пассажиры', economics_data.get('first_class_passengers', 0)),
            ('Coach пассажиры', economics_data.get('coach_passengers', 0)),
            ('Среднее заполнение рейса (%)', f"{economics_data.get('average_load_factor', 0):.2f}%"),
            ('', None),
            ('СРЕДНИЕ ПОКАЗАТЕЛИ', None),
            ('Средний доход на рейс', f"${economics_data.get('average_revenue_per_flight', 0):,.2f}"),
            ('Средняя прибыль на рейс', f"${economics_data.get('average_profit_per_flight', 0):,.2f}"),
        ]
        
        for label, value in financial_metrics:
            if label == '':
                row += 1
                continue
            
            ws[f'A{row}'] = label
            if value is None:
                ws[f'A{row}'].font = Font(bold=True, size=11)
                ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            else:
                ws[f'A{row}'].font = Font(bold=True)
            
            ws[f'B{row}'] = value if value is not None else ''
            if value is not None:
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
    
    @staticmethod
    def _create_monitoring_sheet(ws, data: Dict):
        """Создать лист мониторинга"""
        # Заголовок
        ws['A1'] = "МОНИТОРИНГ РЕСУРСОВ"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Метрики мониторинга
        monitor_metrics = [
            ('ОЧЕРЕДИ', None),
            ('Размер очереди багажа', data.get('luggage_queue_size', 0)),
            ('Размер очереди безопасности', data.get('security_queue_size', 0)),
            ('', None),
            ('ИСПОЛЬЗОВАНИЕ', None),
            ('Использование багажного контроля (%)', f"{data.get('luggage_utilization', 0):.2f}%"),
            ('Использование безопасности (%)', f"{data.get('security_utilization', 0):.2f}%"),
            ('Использование персонала (%)', f"{data.get('staff_utilization', 0):.2f}%"),
            ('Использование багажа (%)', f"{data.get('baggage_utilization', 0):.2f}%"),
            ('Использование терминала (%)', f"{data.get('terminal_utilization', 0):.2f}%"),
            ('', None),
            ('ВРЕМЕННЫЕ ПОКАЗАТЕЛИ', None),
            ('Среднее время ожидания (сек)', round(data.get('avg_wait_time', 0), 2)),
            ('Среднее время задержки (сек)', round(data.get('avg_delay_time', 0), 2)),
            ('Всего задержек', data.get('total_delays', 0)),
        ]
        
        for label, value in monitor_metrics:
            if label == '':
                row += 1
                continue
            
            ws[f'A{row}'] = label
            if value is None:
                ws[f'A{row}'].font = Font(bold=True, size=11)
                ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            else:
                ws[f'A{row}'].font = Font(bold=True)
            
            ws[f'B{row}'] = value if value is not None else ''
            if value is not None:
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
    
    @staticmethod
    def _create_details_sheet(ws, data: Dict):
        """Создать лист деталей"""
        # Заголовок
        ws['A1'] = "ДЕТАЛЬНАЯ ИНФОРМАЦИЯ"
        ws['A1'].font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Вывести все данные в формате ключ-значение
        def write_section(title, section_data):
            nonlocal row
            
            ws[f'A{row}'] = title
            ws[f'A{row}'].font = Font(bold=True, size=11)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            for key, value in section_data.items():
                ws[f'A{row}'] = str(key)
                ws[f'A{row}'].font = Font(size=9)
                
                if isinstance(value, (dict, list)):
                    ws[f'B{row}'] = json.dumps(value, ensure_ascii=False, indent=2)
                else:
                    ws[f'B{row}'] = str(value)
                
                ws[f'B{row}'].alignment = Alignment(horizontal='left', wrap_text=True)
                row += 1
            
            row += 1
        
        # Основная статистика
        main_stats = {k: v for k, v in data.items() 
                     if k not in ['airport_economics', 'terminal_stats', 'active_flights']}
        write_section('Основная статистика', main_stats)
        
        # Терминальная статистика
        if 'terminal_stats' in data:
            write_section('Статистика терминала', data['terminal_stats'])
        
        # Экономика
        if 'airport_economics' in data:
            write_section('Экономические данные', data['airport_economics'])
        
        # Автоматическая ширина колонок
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 50

    def export_pdf(self, stats: Dict, file_path: str) -> bool:
        """
        Экспортировать данные в красивый PDF формате
        
        Args:
            stats: Словарь статистики
            file_path: Путь для сохранения
            
        Returns:
            bool: Успешность операции
        """
        if not REPORTLAB_AVAILABLE:
            print("Ошибка: Требуется установка reportlab (pip install reportlab)")
            return False
        
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, 
                Spacer, PageBreak
            )
            
            prepared_data = self._prepare_data(stats)
            
            # Создать PDF документ
            doc = SimpleDocTemplate(
                file_path,
                pagesize=A4,
                rightMargin=1.5*cm,
                leftMargin=1.5*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            
            # Контейнер для элементов документа
            story = []
            
            # Стили
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor("#1F4E78"),
                spaceAfter=12,
                alignment=1,  # CENTER
                fontName='Helvetica-Bold'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor("#FFFFFF"),
                backColor=colors.HexColor("#1F4E78"),
                spaceAfter=10,
                alignment=0,
                fontName='Helvetica-Bold'
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=11,
                fontName='Helvetica'
            )
            
            # === ТИТУЛЬНАЯ СТРАНИЦА ===
            title = Paragraph("OTCHET O SIMULYACII AEROPORTA", title_style)
            story.append(title)
            
            export_time = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            subtitle = Paragraph(
                f"<font size=12 color='#666666'>Data eksporta: {export_time}</font>",
                normal_style
            )
            story.append(subtitle)
            story.append(Spacer(1, 0.5*cm))
            
            # === ОСНОВНАЯ СТАТИСТИКА ===
            story.append(Paragraph("OSNOVNAYA STATISTIKA", heading_style))
            
            main_data = [
                [Paragraph("Metrika", normal_style), Paragraph("Znachenie", normal_style)],
                [Paragraph("Vsego sobytij", normal_style), Paragraph(str(prepared_data.get('total_events_processed', 0)), normal_style)],
                [Paragraph("Vsego samoletov", normal_style), Paragraph(str(prepared_data.get('total_aircraft', 0)), normal_style)],
                [Paragraph("Vsego passazhirov", normal_style), Paragraph(str(prepared_data.get('total_passengers', 0)), normal_style)],
                [Paragraph("Vremya simulyacii (sec)", normal_style), Paragraph(f"{round(prepared_data.get('simulation_time', 0), 2)}", normal_style)],
                [Paragraph("Propusknaya sposobnost (samoletov/ch)", normal_style), Paragraph(f"{round(prepared_data.get('throughput', 0), 2)}", normal_style)],
                [Paragraph("Ispolzovanie VPP (%)", normal_style), Paragraph(f"{round(prepared_data.get('runway_utilization', 0), 2)}", normal_style)],
                [Paragraph("Ispolzovanie gejatov (%)", normal_style), Paragraph(f"{round(prepared_data.get('gate_utilization', 0), 2)}", normal_style)],
                [Paragraph("Srednee ispolzovanie (%)", normal_style), Paragraph(f"{round(prepared_data.get('average_utilization', 0), 2)}", normal_style)],
            ]
            
            main_table = Table(main_data, colWidths=[8*cm, 5*cm])
            main_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#CCCCCC")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            ]))
            story.append(main_table)
            story.append(Spacer(1, 0.5*cm))
            
            # === ЭКОНОМИЧЕСКАЯ СТАТИСТИКА ===
            if 'airport_economics' in prepared_data:
                story.append(PageBreak())
                story.append(Paragraph("EKONOMICHESKAYA STATISTIKA", heading_style))
                
                economics = prepared_data['airport_economics']
                econ_data = [
                    [Paragraph("Pokazatel", normal_style), Paragraph("Znachenie", normal_style)],
                    [Paragraph("Vsego rejsov", normal_style), Paragraph(str(economics.get('total_flights', 0)), normal_style)],
                    [Paragraph("Mestnye rejsy", normal_style), Paragraph(str(economics.get('commuter_flights', 0)), normal_style)],
                    [Paragraph("Mezhdunarodnye rejsy", normal_style), Paragraph(str(economics.get('international_flights', 0)), normal_style)],
                    [Paragraph("Obschij dokhod", normal_style), Paragraph(f"${economics.get('total_revenue', 0):,.2f}", normal_style)],
                    [Paragraph("Dokhod First Class", normal_style), Paragraph(f"${economics.get('first_class_revenue', 0):,.2f}", normal_style)],
                    [Paragraph("Dokhod Coach", normal_style), Paragraph(f"${economics.get('coach_revenue', 0):,.2f}", normal_style)],
                    [Paragraph("Obschie raskhody", normal_style), Paragraph(f"${economics.get('total_costs', 0):,.2f}", normal_style)],
                    [Paragraph("Obschaya pribyl", normal_style), Paragraph(f"${economics.get('total_profit', 0):,.2f}", normal_style)],
                    [Paragraph("ROI (%)", normal_style), Paragraph(f"{economics.get('roi_percentage', 0):.2f}%", normal_style)],
                    [Paragraph("Vsego passazhirov", normal_style), Paragraph(str(economics.get('total_passengers_served', 0)), normal_style)],
                    [Paragraph("Srednee zapolnenie (%)", normal_style), Paragraph(f"{economics.get('average_load_factor', 0):.2f}%", normal_style)],
                ]
                
                econ_table = Table(econ_data, colWidths=[8*cm, 5*cm])
                econ_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F0F8FF")),
                    ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#CCCCCC")),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ]))
                story.append(econ_table)
                story.append(Spacer(1, 0.5*cm))
            
            # === МОНИТОРИНГ РЕСУРСОВ ===
            story.append(PageBreak())
            story.append(Paragraph("MONITORING RESURSOV", heading_style))
            
            monitor_data = [
                [Paragraph("Resurs", normal_style), Paragraph("Pokazatel", normal_style)],
                [Paragraph("Ochered' bagazha", normal_style), Paragraph(str(prepared_data.get('luggage_queue_size', 0)), normal_style)],
                [Paragraph("Ochered' bezopasnosti", normal_style), Paragraph(str(prepared_data.get('security_queue_size', 0)), normal_style)],
                [Paragraph("Ispolzovanie bagazhnogo kontrolya (%)", normal_style), Paragraph(f"{prepared_data.get('luggage_utilization', 0):.2f}", normal_style)],
                [Paragraph("Ispolzovanie bezopasnosti (%)", normal_style), Paragraph(f"{prepared_data.get('security_utilization', 0):.2f}", normal_style)],
                [Paragraph("Ispolzovanie personala (%)", normal_style), Paragraph(f"{prepared_data.get('staff_utilization', 0):.2f}", normal_style)],
                [Paragraph("Srednee vremya ozhidaniya (sec)", normal_style), Paragraph(f"{round(prepared_data.get('avg_wait_time', 0), 2)}", normal_style)],
                [Paragraph("Srednee vremya zaderzhki (sec)", normal_style), Paragraph(f"{round(prepared_data.get('avg_delay_time', 0), 2)}", normal_style)],
                [Paragraph("Vsego zaderzhek", normal_style), Paragraph(str(prepared_data.get('total_delays', 0)), normal_style)],
            ]
            
            monitor_table = Table(monitor_data, colWidths=[8*cm, 5*cm])
            monitor_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#FFF8DC")),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#CCCCCC")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            ]))
            story.append(monitor_table)
            
            # === НИЖНИЙ КОЛОНТИТУЛ ===
            story.append(Spacer(1, 1*cm))
            footer = Paragraph(
                f"<font size=9 color='#999999'>AeroSim EDU - Airport Simulation - {export_time}</font>",
                normal_style
            )
            story.append(footer)
            
            # Построить PDF
            doc.build(story)
            return True
            
        except Exception as e:
            print(f"Oshibka pri eksporte PDF: {e}")
            import traceback
            traceback.print_exc()
            return False

